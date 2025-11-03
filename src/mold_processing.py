import pdfplumber
from openpyxl.styles import Font, Alignment
import math

#CONSTNATS
LAB_REFERENCE_NUMBER_STYLE = Font(name='Arial', size=11, bold=True)
OTHER_STYLE = Font(name='Arial', size=11)

def find_mold_values(pdf_path):
    """
    Extracts mold types and their corresponding values from the 'Outdoor' section
    of the second page of a PDF.

    Args:
        pdf_path (str): Path to the PDF file.

    Returns:
        tuple: (mold_dict, lab_reference_number)
            mold_dict (dict): Dictionary mapping mold types to their values.
            lab_reference_number (str): The lab reference number found in the table.
        or (None, None) if 'Outdoor' section is not found.
    """
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[1]  # 0-based index, so 1 is the second page
        tables = page.extract_tables()
        for table_num, table in enumerate(tables):
            transposed = list(zip(*table))
            outdoor_col_index = None
            outdoor_row_index = None
            for col_idx, col in enumerate(transposed):
                for row_idx, cell in enumerate(col):
                    if cell and (cell.strip().lower() == "outdoor" or cell.strip() == "outdoors" or cell.strip() == "extérieur"):
                        outdoor_col_index = col_idx
                        outdoor_row_index = row_idx
                        break
                if outdoor_col_index is not None:
                    break
            if outdoor_col_index is not None:
                outdoor_column = transposed[outdoor_col_index]
                lab_reference_number = transposed[outdoor_col_index][1]
                mold_col_index = outdoor_col_index + 2
                mold_types = list(mt.strip().replace(",","") if mt else "" for mt in transposed[0][3:])
                mold_values = list(transposed[mold_col_index][3:])
                mold_dict = {}
                for mold_type, value in zip(mold_types, mold_values):
                    if mold_type and mold_type.strip():
                        cleaned = value.strip().replace(",", "") if value else ""
                    mold_dict[mold_type.strip()] = int(cleaned) if (cleaned and cleaned.isdigit()) else None
                info = (mold_dict, lab_reference_number)
            else:
                print("'Outdoor' not found in this table.")
                mold_dict = None
                lab_reference_number = None
                info = (mold_dict, lab_reference_number)
    return info

def find_total_count_index(sheet):
    """
    Finds the index of the 'Total' column in the Excel sheet.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        int: The 0-based index of the 'Total' column.

    Raises:
        ValueError: If the 'Total' column header is missing in the Excel sheet.
    """
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        total_col_index = header_row.index("Total")
    except ValueError:
        raise ValueError("The 'Total' column header is missing in the Excel sheet.")
    return total_col_index

def insert_into_excel(mold_dict, sheet, lab_reference_number):
    """
    Inserts mold counts into the first empty column of an Excel sheet, using the lab reference number as the header.

    Args:
        mold_dict (dict): Dictionary mapping mold types to their values.
        sheet (Worksheet): The active worksheet.
        lab_reference_number (str): The lab reference number to use as the column header.

    Returns:
        None
    """
    total_count_index = find_total_count_index(sheet)
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    for col_index in range(total_count_index):
        if header_row[col_index] is None:
            break
    else:
        col_index = total_count_index
        sheet.insert_cols(col_index + 1)
    new_col_header = lab_reference_number
    sheet.cell(row=3, column=col_index + 1, value=new_col_header).font = LAB_REFERENCE_NUMBER_STYLE
    for row in range(4, sheet.max_row + 1):
        spore_type = str(sheet.cell(row=row, column=1).value).strip()
        if spore_type in mold_dict:
            sheet.cell(row=row, column=col_index + 1, value=mold_dict[spore_type]).font = OTHER_STYLE

    return

def total_count(sheet):
    """
    Calculates and writes the sum of mold counts for each row into the 'Total' column of the Excel sheet.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    """
    total_count_index = find_total_count_index(sheet)
    for row_idx in range(4, sheet.max_row + 1):
        count = 0
        for col_idx in range(2, total_count_index + 1):  # Assuming data starts at column 2 (B)
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                count += int(cell_value)
        sheet.cell(row=row_idx, column=total_count_index + 1).value = count
    return
def clear_old_stats(sheet):
    """Clears previous Min/Percentile/Median/Max/Stdv values from the sheet."""
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    for row in range(4, sheet.max_row + 1):
        for col_name in ["Min", "5th Percentile", "Median", "95th Percentile", "Max", "Stdv"]:
            try:
                col_idx = header_row.index(col_name) + 1
                sheet.cell(row=row, column=col_idx, value=0)
            except ValueError:
                pass

def mean_count(sheet):
    """
    Calculates and writes the mean of mold counts for each row into a new column labeled 'Mean' in the Excel sheet.
    Blank cells are ignored in the mean calculation.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    """
    total_count_index = find_total_count_index(sheet)
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        mean_col_index = header_row.index("Mean")
    except ValueError:
        mean_col_index = total_count_index + 1
        sheet.insert_cols(mean_col_index + 1)
        sheet.cell(row=3, column=mean_col_index + 1, value="Mean").font = LAB_REFERENCE_NUMBER_STYLE
    for row_idx in range(4, sheet.max_row + 1):
        values = []
        for col_idx in range(2, total_count_index + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                values.append(int(cell_value))
        mean = sum(values) / len(values) if values else None
        sheet.cell(row=row_idx, column=mean_col_index + 1).value = mean
    return

def stdv_count(sheet):
    """
    Calculates and writes the standard deviation of mold counts for each row into a new column labeled 'Stdv' in the Excel sheet.
    Blank cells are ignored in the standard deviation calculation.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    """
    total_count_index = find_total_count_index(sheet)
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        stdv_col_index = header_row.index("Stdv")
    except ValueError:
        try:
            mean_col_index = header_row.index("Mean")
            stdv_col_index = mean_col_index + 1
        except ValueError:
            stdv_col_index = total_count_index + 2
        sheet.insert_cols(stdv_col_index + 1)
        sheet.cell(row=3, column=stdv_col_index + 1, value="Stdv").font = LAB_REFERENCE_NUMBER_STYLE
    for row_idx in range(4, sheet.max_row + 1):
        values = []
        for col_idx in range(2, total_count_index + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                values.append(int(cell_value))
        stdv = None
        if values and len(values) > 1:
            mean = sum(values) / len(values)
            variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
            stdv = math.sqrt(variance)
        sheet.cell(row=row_idx, column=stdv_col_index + 1).value = stdv
    return

def display_mold_type_frequency(sheet):
    """
    Calculates and writes the frequency (number of samples where each mold type appears)
    into a new column labeled 'Frequency' in the Excel sheet.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    """
    # Find or create the "Frequency" column (after "Stdv" if present, else at the end)
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        freq_col_index = header_row.index("Frequency")
    except ValueError:
        try:
            stdv_col_index = header_row.index("Stdv")
            freq_col_index = stdv_col_index + 1
        except ValueError:
            freq_col_index = len(header_row)
        sheet.insert_cols(freq_col_index + 1)
        sheet.cell(row=3, column=freq_col_index + 1, value="Frequency").font = LAB_REFERENCE_NUMBER_STYLE

    # For each row (mold type), count the number of nonzero, non-blank cells (excluding column 1)

    total_col_index = header_row.index("Total")
    num_samples = total_col_index - 1  # Number of sample columns (excluding mold type name)
    for row in range(4, sheet.max_row + 1):
        frequency = 0
        for col in range(2, total_col_index + 1):  # Skip column 1 (mold type name)
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value not in (None, 0, "", "0"):
                frequency += 1
        percent = (frequency / num_samples) * 100 if num_samples > 0 else 0
        if isinstance(percent, float):
            percent = round(percent, 2)
        sheet.cell(row=row, column=freq_col_index + 1, value=percent)
    return

def find_min(sheet):
    """
    Finds the minimum spore count of each spore

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        int: The minimum value found in the 'Total' column.
    """
    #Find or create the "Min" column (after "Frequency" if present, else at the end)
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        min_col_index = header_row.index("Min")
    except ValueError:
        try:
            freq_col_index = header_row.index("Frequency")
            min_col_index = freq_col_index + 1
        except ValueError:
            min_col_index = len(header_row)
        sheet.insert_cols(min_col_index + 1)
        sheet.cell(row=3, column=min_col_index + 1, value="Min").font = LAB_REFERENCE_NUMBER_STYLE

    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    total_col_index = header_row.index("Total")
    for row in range(4, sheet.max_row + 1):
        min_value = float('inf')
        found = False
        for col in range(2, total_col_index):  # Sample columns only
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                cell_value = int(cell_value)  # Ensure cell_value is an integer
                if cell_value < min_value:
                    min_value = cell_value
                found = True
        sheet.cell(row=row, column=min_col_index + 1, value=min_value if found else None)

    return

def fifth_percentile(sheet):
    """ 
    Calculates and writes the 5th percentile of mold counts for each row into a new column labeled '5th Percentile' in the Excel sheet.
    Blank and zero cells are ignored in the calculation.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    
    """
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        fifth_percentile_col_index = header_row.index("5th Percentile")
    except ValueError:
        try:
            min_col_index = header_row.index("Min")
            fifth_percentile_col_index = min_col_index + 1
        except ValueError:
            fifth_percentile_col_index = len(header_row)
        sheet.insert_cols(fifth_percentile_col_index + 1)
        sheet.cell(row=3, column=fifth_percentile_col_index + 1, value="5th Percentile").font = LAB_REFERENCE_NUMBER_STYLE
    
    total_col_index = header_row.index("Total")
    for row in range(4, sheet.max_row + 1):
        values = []
        # Collect all sample values for this row (columns 2 up to "Total")
        for col in range(2, total_col_index):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                values.append(int(cell_value))
        if values:
            values.sort()
            # Nearest-rank method for 5th percentile
            k = max(0, int(math.ceil(0.05 * len(values))) - 1)
            percentile_value = values[k]
        else:
            percentile_value = None
        sheet.cell(row=row, column=fifth_percentile_col_index + 1, value=percentile_value)
    return

def find_median(sheet):
    """
    Calculates and writes the median of mold counts for each row into a new column labeled 'Median' in the Excel sheet.
    Blank and zero cells are ignored in the calculation.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    """
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        median_col_index = header_row.index("Median")
    except ValueError:
        try:
            fifth_percentile_col_index = header_row.index("5th Percentile")
            median_col_index = fifth_percentile_col_index + 1
        except ValueError:
            median_col_index = len(header_row)
        sheet.insert_cols(median_col_index + 1)
        sheet.cell(row=3, column=median_col_index + 1, value="Median").font = LAB_REFERENCE_NUMBER_STYLE
    
    total_col_index = header_row.index("Total")
    for row in range(4, sheet.max_row + 1):
        values = []
        # Collect all sample values for this row (columns 2 up to "Total")
        for col in range(2, total_col_index):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                values.append(int(cell_value))
        if values:
            values.sort()
            n = len(values)
            if n % 2 == 0:  # Even number of elements
                median_value = (values[n // 2 - 1] + values[n // 2]) / 2
            else:  # Odd number of elements
                median_value = values[n // 2]
        else:
            median_value = None
        sheet.cell(row=row, column=median_col_index + 1, value=median_value)
    return

def find_ninety_fifth_percentile(sheet):
    """
    
    Calculates and writes the 95th percentile of mold counts for each row into a new column labeled '95th Percentile' in the Excel sheet.
    Blank and zero cells are ignored in the calculation.

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        None
    
    """
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        ninty_fifth_percentile_col_index = header_row.index("95th Percentile")
    except ValueError:
        try:
            min_col_index = header_row.index("Median")
            ninty_fifth_percentile_col_index = min_col_index + 1
        except ValueError:
            ninty_fifth_percentile_col_index = len(header_row)
        sheet.insert_cols(ninty_fifth_percentile_col_index + 1)
        sheet.cell(row=3, column=ninty_fifth_percentile_col_index + 1, value="95th Percentile").font = LAB_REFERENCE_NUMBER_STYLE
    
    total_col_index = header_row.index("Total")
    for row in range(4, sheet.max_row + 1):
        values = []
        # Collect all sample values for this row (columns 2 up to "Total")
        for col in range(2, total_col_index):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                values.append(int(cell_value))
        if values:
            values.sort()
            # Nearest-rank method for 95th percentile
            k = max(0, int(math.ceil(0.95 * len(values))) - 1)
            percentile_value = values[k]
        else:
            percentile_value = None
        sheet.cell(row=row, column=ninty_fifth_percentile_col_index + 1, value=percentile_value)
    return

def find_max(sheet):
    """
    Finds the maximum spore count of each spore

    Args:
        sheet (Worksheet): The active worksheet.

    Returns:
        int: The maximum value found in the 'Total' column.
    """
    #Find or create the "Max" column (after "Frequency" if present, else at the end)
   
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        max_col_index = header_row.index("Max")
    except ValueError:
        try:
            prev_col_index = header_row.index("95th Percentile")
            max_col_index = prev_col_index + 1
        except ValueError:
            max_col_index = len(header_row)
        sheet.insert_cols(max_col_index + 1)
        sheet.cell(row=3, column=max_col_index + 1, value="Max").font = LAB_REFERENCE_NUMBER_STYLE

    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    total_col_index = header_row.index("Total")
    for row in range(4, sheet.max_row + 1):
        max_value = float('-inf')
        found = False
        for col in range(2, total_col_index):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                cell_value = int(cell_value)  # Ensure cell_value is an integer
                if cell_value > max_value:
                    max_value = cell_value
                found = True
        sheet.cell(row=row, column=max_col_index + 1, value=max_value if found else None)
    return

def find_count(sheet):
    """
    Counts the number of non-blank cells in each row and writes the count into a new column labeled 'Count'.

    Args:
        sheet (Worksheet): The active worksheet.
    
    Returns:
        None
    """
    header_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    try:
        count_col_index = header_row.index("Count")
    except ValueError:
        try:
            max_col_index = header_row.index("Max")
            count_col_index = max_col_index + 1
        except ValueError:
            count_col_index = len(header_row)
        sheet.insert_cols(count_col_index + 1)
        sheet.cell(row=3, column=count_col_index + 1, value="Count").font = LAB_REFERENCE_NUMBER_STYLE
    
    total_col_index = find_total_count_index(sheet)
    for row in range(4, sheet.max_row + 1):
        count = 0
        for col in range(2, total_col_index):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                count += 1
        sheet.cell(row=row, column=count_col_index + 1, value=count)